import streamlit as st
import requests
import pandas as pd
import json, time, io, hashlib, math, unicodedata
from datetime import datetime

st.set_page_config(page_title="Solar Lead Finder", page_icon="☀️", layout="wide", initial_sidebar_state="expanded")

# ── CREDENCIAIS ──────────────────────────────────────────────────────────────
USUARIOS = {
    "admin":  hashlib.sha256("solar2024".encode()).hexdigest(),
    "equipe": hashlib.sha256("leads2024".encode()).hexdigest(),
}
def hash_senha(s): return hashlib.sha256(s.encode()).hexdigest()

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""<style>
.stApp{background:#0f1117}
section[data-testid="stSidebar"]{background:#161b27;border-right:1px solid #2d3748}
h1,h2,h3,h4{color:#f1f5f9!important}
p,label{color:#e2e8f0!important}
.stMarkdown p{color:#94a3b8!important}
.stSelectbox>div,.stMultiSelect>div,.stTextInput>div{background:#1e293b!important;border:1px solid #2d3748!important;border-radius:8px!important;color:#e2e8f0!important}
.stButton>button{background:#f59e0b!important;color:#1c1400!important;border:none!important;border-radius:8px!important;font-weight:600!important;width:100%!important;padding:12px!important;font-size:14px!important}
.stButton>button:hover{background:#fbbf24!important}
.stDownloadButton>button{background:#1e3a5f!important;color:#60a5fa!important;border:1px solid #1d4ed8!important;border-radius:8px!important;width:100%!important}
[data-testid="stMetric"]{background:#161b27!important;border:1px solid #2d3748!important;border-radius:10px!important;padding:16px!important}
[data-testid="stMetricValue"]{color:#f1f5f9!important;font-size:28px!important}
[data-testid="stMetricLabel"]{color:#475569!important;font-size:12px!important}
hr{border-color:#2d3748!important}
.stCheckbox label{color:#cbd5e1!important;font-size:13px!important}
.badge-src{display:inline-block;background:#052e16;color:#4ade80;border:1px solid #166534;border-radius:4px;padding:2px 8px;font-size:11px;margin-right:4px}
</style>""", unsafe_allow_html=True)

# ── LOGIN ─────────────────────────────────────────────────────────────────────
def tela_login():
    c1,c2,c3 = st.columns([1,1.2,1])
    with c2:
        st.markdown("<br><br>",unsafe_allow_html=True)
        st.markdown('<div style="background:#161b27;border:0.5px solid #2d3748;border-radius:14px;padding:36px 32px;text-align:center"><div style="width:52px;height:52px;background:#f59e0b;border-radius:12px;display:inline-flex;align-items:center;justify-content:center;font-size:26px;margin-bottom:16px">☀️</div><h2 style="color:#f1f5f9;margin:0 0 4px">Solar Lead Finder</h2><p style="color:#475569;font-size:13px">Acesso restrito</p></div>',unsafe_allow_html=True)
        st.markdown("<br>",unsafe_allow_html=True)
        u = st.text_input("Usuário", placeholder="usuário")
        s = st.text_input("Senha", type="password", placeholder="••••••••")
        if st.button("Entrar"):
            if u in USUARIOS and USUARIOS[u] == hash_senha(s):
                st.session_state.autenticado = True
                st.session_state.usuario = u
                st.rerun()
            else:
                st.error("Usuário ou senha incorretos.")

if "autenticado" not in st.session_state: st.session_state.autenticado = False
if not st.session_state.autenticado:
    tela_login()
    st.stop()

# ── COORDENADAS DOS MUNICÍPIOS (sem depender de API externa) ──────────────────
COORDS = {
    "TEÓFILO OTONI":(-17.8578,-41.5053),"ARAÇUAÍ":(-16.8519,-42.0706),
    "DIAMANTINA":(-18.2428,-43.6003),"ITAMARANDIBA":(-17.8558,-42.8564),
    "MONTES CLAROS":(-16.7286,-43.8611),"JANUÁRIA":(-15.4878,-44.3619),
    "UNAÍ":(-16.3597,-46.9044),"CARATINGA":(-19.7892,-42.1392),
    "IPATINGA":(-19.4686,-42.5375),"GOVERNADOR VALADARES":(-18.8536,-41.9497),
    "UBERLÂNDIA":(-18.9186,-48.2772),"BELO HORIZONTE":(-19.9167,-43.9345),
    "PEDRA AZUL":(-16.0064,-41.2733),"CAPELINHA":(-17.6947,-42.5167),
    "PORTEIRINHA":(-15.7436,-43.0258),"NANUQUE":(-17.8394,-40.3531),
    "BOCAIÚVA":(-17.1122,-43.8142),"PIRAPORA":(-17.3447,-44.9417),
    "SÃO FRANCISCO":(-15.9500,-44.8639),"JANAÚBA":(-15.8028,-43.3086),
    "VIRGEM DA LAPA":(-16.9822,-42.3481),"CORONEL MURTA":(-16.6156,-42.1917),
    "TURMALINA":(-17.2814,-42.7258),"MINAS NOVAS":(-17.2269,-42.5664),
    "CARLOS CHAGAS":(-17.7028,-40.7597),"POTÉ":(-17.8122,-41.7853),
    "MALACACHETA":(-17.8383,-42.0700),"ITAMBACURI":(-18.0333,-41.6844),
    "ATALÉIA":(-18.0483,-41.1214),"LADAINHA":(-17.6264,-41.7403),
    "JUIZ DE FORA":(-21.7642,-43.3503),"VIÇOSA":(-20.7561,-42.8828),
    "MANHUAÇU":(-20.2578,-42.0281),"UBERLÂNDIA":(-18.9186,-48.2772),
    "UBERABA":(-19.7486,-47.9322),"PATOS DE MINAS":(-18.5786,-46.5183),
    "DIVINÓPOLIS":(-20.1386,-44.8853),"CONTAGEM":(-19.9317,-44.0536),
    "SETE LAGOAS":(-19.4658,-44.2475),"VARGINHA":(-21.5511,-45.4306),
    "POÇOS DE CALDAS":(-21.7872,-46.5617),"POUSO ALEGRE":(-22.2300,-45.9358),
    "BARREIRAS":(-12.1522,-44.9906),"LUÍS EDUARDO MAGALHÃES":(-12.0964,-45.7939),
    "MUCUGÊ":(-13.0008,-41.3625),"LENÇÓIS":(-12.5603,-41.3886),
    "JACOBINA":(-11.1817,-40.5133),"IRECÊ":(-11.3036,-41.8550),
    "JUAZEIRO":(-9.4278,-40.5028),"PAULO AFONSO":(-9.4042,-38.2172),
    "SENHOR DO BONFIM":(-10.4628,-40.1886),"SALVADOR":(-12.9714,-38.5014),
    "CAMAÇARI":(-12.6997,-38.3242),"FEIRA DE SANTANA":(-12.2664,-38.9663),
    "VITÓRIA DA CONQUISTA":(-14.8661,-40.8444),"ILHÉUS":(-14.7881,-39.0494),
    "ITABUNA":(-14.7853,-39.2800),"CORRENTINA":(-13.3417,-44.6408),
    "SÃO DESIDÉRIO":(-12.3636,-44.9750),"BOM JESUS DA LAPA":(-13.2558,-43.4183),
    "VITÓRIA":(-20.3155,-40.3128),"VILA VELHA":(-20.3297,-40.2922),
    "SERRA":(-20.1286,-40.3075),"CACHOEIRO DE ITAPEMIRIM":(-20.8486,-41.1133),
    "LINHARES":(-19.3917,-40.0631),"SÃO MATEUS":(-18.7153,-39.8589),
    "COLATINA":(-19.5386,-40.6275),"ALEGRE":(-20.7622,-41.5331),
    "ARACRUZ":(-19.8197,-40.2742),"NOVA VENÉCIA":(-18.7153,-40.4044),
    "GOIÂNIA":(-16.6864,-49.2643),"ANÁPOLIS":(-16.3281,-48.9531),
    "RIO VERDE":(-17.7983,-50.9281),"CUIABÁ":(-15.6014,-56.0979),
    "SINOP":(-11.8642,-55.5022),"SÃO PAULO":(-23.5505,-46.6333),
    "BELÉM":(-1.4558,-48.5044),"MARABÁ":(-5.3686,-49.1178),
    "PARAUAPEBAS":(-6.0686,-49.9019),"SOROCABA":(-23.5015,-47.4526),
}

def normalizar(s):
    return ''.join(c for c in unicodedata.normalize('NFD', s.upper().strip()) if unicodedata.category(c) != 'Mn')

def get_coords_municipio(municipio):
    key = municipio.upper().strip()
    if key in COORDS: return COORDS[key]
    key_n = normalizar(key)
    for k,v in COORDS.items():
        if normalizar(k) == key_n: return v
    # Fallback: Nominatim
    try:
        r = requests.get("https://nominatim.openstreetmap.org/search",
                         params={"q":f"{municipio}, Brasil","format":"json","limit":1},
                         headers={"User-Agent":"SolarLeadFinder/2.0"}, timeout=8)
        if r.status_code == 200:
            d = r.json()
            if d: return float(d[0]["lat"]), float(d[0]["lon"])
    except: pass
    return None, None

def distancia_km(lat1,lon1,lat2,lon2):
    try:
        R=6371; dl=math.radians(float(lat2)-float(lat1)); dlo=math.radians(float(lon2)-float(lon1))
        a=math.sin(dl/2)**2+math.cos(math.radians(float(lat1)))*math.cos(math.radians(float(lat2)))*math.sin(dlo/2)**2
        return R*2*math.atan2(math.sqrt(a),math.sqrt(1-a))
    except: return 9999

# ── REGIÕES ───────────────────────────────────────────────────────────────────
REGIOES = {
    "MG":{"Todas as regiões":[],"Vale do Jequitinhonha":["Araçuaí","Diamantina","Itamarandiba","Pedra Azul","Turmalina","Capelinha","Minas Novas","Virgem da Lapa","Coronel Murta","Berilo"],"Norte de Minas":["Montes Claros","Januária","Pirapora","Bocaiúva","Janaúba","Porteirinha","Unaí","São Francisco","Brasília de Minas"],"Vale do Mucuri":["Teófilo Otoni","Nanuque","Ladainha","Carlos Chagas","Novo Oriente de Minas","Poté","Malacacheta","Itambacuri","Ataléia","Franciscópolis"],"Zona da Mata":["Juiz de Fora","Viçosa","Caratinga","Governador Valadares","Manhuaçu","Ipatinga"],"Triângulo Mineiro":["Uberlândia","Uberaba","Araguari","Ituiutaba","Patos de Minas"],"Sul de Minas":["Poços de Caldas","Varginha","Pouso Alegre","Itajubá","Lavras"],"Central":["Belo Horizonte","Contagem","Betim","Sete Lagoas","Divinópolis"]},
    "BA":{"Todas as regiões":[],"Oeste Baiano":["Barreiras","Luís Eduardo Magalhães","Bom Jesus da Lapa","Correntina","São Desidério"],"Chapada Diamantina":["Mucugê","Lençóis","Palmeiras","Seabra","Andaraí","Piatã"],"Sertão Baiano":["Juazeiro","Paulo Afonso","Senhor do Bonfim","Jacobina","Irecê","Xique-Xique"],"Litoral/Metropolitana":["Salvador","Camaçari","Feira de Santana","Lauro de Freitas"],"Sul da Bahia":["Ilhéus","Itabuna","Eunápolis","Porto Seguro","Vitória da Conquista"]},
    "ES":{"Todas as regiões":[],"Norte do ES":["São Mateus","Linhares","Colatina","Nova Venécia"],"Metropolitana":["Vitória","Vila Velha","Serra","Cariacica","Guarapari","Aracruz"],"Sul do ES":["Cachoeiro de Itapemirim","Alegre","Guaçuí","Castelo"]},
    "GO":{"Todas as regiões":[],"Sul de Goiás":["Rio Verde","Jataí","Itumbiara","Caldas Novas"],"Centro":["Goiânia","Anápolis"]},
    "MT":{"Todas as regiões":[],"Norte":["Sinop","Alta Floresta","Sorriso","Lucas do Rio Verde"],"Sul":["Cuiabá","Rondonópolis"]},
    "PA":{"Todas as regiões":[],"Pará":["Belém","Santarém","Marabá","Parauapebas"]},
}
TODOS_ESTADOS = ["MG","BA","ES","SP","GO","MT","PA","RO","TO","MS","PR","RS","SC","RJ","PE","CE","PI","MA","RN","PB","AL","SE","AM"]

# ── DADOS DEMO ────────────────────────────────────────────────────────────────
def gerar_demo(estado):
    base = {
        "MG":[
            ("Rural-Irrigacao","Fazenda Serra Verde","Araçuaí",480,-16.85,-42.07,"Alto","Sem rede próxima","Área: 480 ha · Outorga ANA","","",""),
            ("Rural-Irrigacao","Fazenda Três Barras","Januária",620,-15.48,-44.36,"Alto","Sem rede próxima","Área: 620 ha · Pivô central","","",""),
            ("Rural-Irrigacao","Agropecuária Jequitinhonha","Diamantina",340,-18.24,-43.60,"Alto","Sem rede próxima","Área: 340 ha","","",""),
            ("Rural-Irrigacao","Fazenda do Cerrado","Unaí",890,-16.36,-46.90,"Alto","Sem rede próxima","Área: 890 ha · Pivô central","","",""),
            ("Rural-Irrigacao","Sítio Boa Esperança","Montes Claros",210,-16.72,-43.86,"Médio","Verificar","Área: 210 ha","","",""),
            ("Rural-Irrigacao","Fazenda Santa Luzia","Porteirinha",430,-15.74,-43.02,"Alto","Sem rede próxima","Área: 430 ha","","",""),
            ("Rural-Irrigacao","Grupo Rural Mucuri","Teófilo Otoni",180,-17.86,-41.50,"Médio","Verificar","Área: 180 ha","","",""),
            ("Rural-Irrigacao","Fazenda Vale Verde","Nanuque",260,-17.83,-40.35,"Médio","Sem rede próxima","Área: 260 ha","","",""),
            ("Mineracao","Min. Vale do Jequitinhonha","Itamarandiba",320,-17.85,-42.85,"Alto","Área remota — alto risco diesel","Concessão de Lavra · Ferro","","",""),
            ("Mineracao","Mineradora São Domingos","Diamantina",520,-18.24,-43.60,"Alto","Área remota — alto risco diesel","Concessão de Lavra · Ouro","","",""),
            ("Mineracao","Granitos MG Extração","Pedra Azul",85,-16.00,-41.27,"Médio","Área remota — alto risco diesel","Licenciamento · Granito","","",""),
            ("Mineracao","Mineração Capelinha","Capelinha",210,-17.69,-42.52,"Alto","Área remota — alto risco diesel","Concessão de Lavra · Nióbio","","",""),
            ("Mineracao","Extratora Norte MG","Bocaiúva",145,-17.11,-43.81,"Médio","Área remota — alto risco diesel","Licenciamento · Calcário","","",""),
            ("ACL-GrupoA","Cerâmica Industrial Mucuri SA","Teófilo Otoni",0,-17.86,-41.50,"Alto","Conectado — Grupo A / Mercado Livre","Demanda: 2.4 MW","12.345.678/0001-90","João Silva","33991234567"),
            ("ACL-GrupoA","Siderúrgica Vale do Aço Ltda","Ipatinga",0,-19.47,-42.54,"Alto","Conectado — Grupo A / Mercado Livre","Demanda: 8.1 MW","23.456.789/0001-11","Maria Santos","31991234567"),
            ("ACL-GrupoA","Frigorífico Central MG SA","Uberlândia",0,-18.91,-48.27,"Alto","Conectado — Grupo A / Mercado Livre","Demanda: 1.2 MW","34.567.890/0001-22","Pedro Alves","34991234567"),
            ("ACL-GrupoA","Cimento Norte MG Ltda","Montes Claros",0,-16.72,-43.86,"Alto","Conectado — Grupo A / Mercado Livre","Demanda: 5.5 MW","45.678.901/0001-33","Ana Lima","38991234567"),
            ("Diesel-Gerador","Frigorífico Bom Pastor Ltda","Caratinga",0,-19.78,-42.13,"Médio","Usa diesel — candidato solar/híbrido","Potência: 320 kW","","",""),
            ("Diesel-Gerador","Mineração São Domingos Ltda","Diamantina",0,-18.24,-43.60,"Alto","Usa diesel — candidato solar/híbrido","Potência: 850 kW","","",""),
            ("Diesel-Gerador","Pedreiras Jequitinhonha","Araçuaí",0,-16.85,-42.07,"Alto","Usa diesel — candidato solar/híbrido","Potência: 560 kW","","",""),
            ("Diesel-Gerador","Cerâmica Mucuri Ltda","Teófilo Otoni",0,-17.86,-41.50,"Médio","Usa diesel — candidato solar/híbrido","Potência: 240 kW","","",""),
            ("Diesel-Gerador","Frigorifico Norte MG","Montes Claros",0,-16.72,-43.86,"Médio","Usa diesel — candidato solar/híbrido","Potência: 180 kW","","",""),
        ],
        "BA":[
            ("Rural-Irrigacao","Fazenda Chapada","Barreiras",720,-12.15,-44.99,"Alto","Sem rede próxima","Área: 720 ha · Pivô central","","",""),
            ("Rural-Irrigacao","Agropecuária São João","Luís Eduardo Magalhães",340,-12.09,-45.79,"Alto","Sem rede próxima","Área: 340 ha · Pivô central","","",""),
            ("Rural-Irrigacao","Grupo Rural Oeste","Correntina",810,-13.34,-44.64,"Alto","Sem rede próxima","Área: 810 ha · Irrigação por pivô","","",""),
            ("Rural-Irrigacao","Fazenda Bom Sucesso","São Desidério",550,-12.36,-44.97,"Alto","Sem rede próxima","Área: 550 ha","","",""),
            ("Rural-Irrigacao","Fazenda Sertão Verde","Irecê",290,-11.30,-41.85,"Médio","Verificar","Área: 290 ha","","",""),
            ("Mineracao","Mineradora Chapada Diamantina","Mucugê",560,-13.00,-41.36,"Alto","Área remota — alto risco diesel","Concessão de Lavra · Diamante","","",""),
            ("Mineracao","Extratora Bahia Ltda","Jacobina",190,-11.18,-40.51,"Médio","Área remota — alto risco diesel","Licenciamento · Ouro","","",""),
            ("Mineracao","Min. Sertão Baiano","Senhor do Bonfim",280,-10.46,-40.19,"Alto","Área remota — alto risco diesel","Concessão de Lavra · Cromo","","",""),
            ("ACL-GrupoA","Petroquímica Camaçari Ltda","Camaçari",0,-12.69,-38.32,"Alto","Conectado — Grupo A / Mercado Livre","Demanda: 15.0 MW","45.678.901/0001-44","Maria Souza","71991234567"),
            ("ACL-GrupoA","Cerâmica Oeste Baiano","Barreiras",0,-12.15,-44.99,"Alto","Conectado — Grupo A / Mercado Livre","Demanda: 3.2 MW","56.789.012/0001-55","Carlos Lima","77991234567"),
            ("Diesel-Gerador","Mineradora Chapada Diesel","Mucugê",0,-13.00,-41.36,"Alto","Usa diesel — candidato solar/híbrido","Potência: 720 kW","","",""),
            ("Diesel-Gerador","Cerâmica Nordeste SA","Vitória da Conquista",0,-14.86,-40.84,"Médio","Usa diesel — candidato solar/híbrido","Potência: 430 kW","","",""),
            ("Diesel-Gerador","Fazenda Agropastoril","Irecê",0,-11.30,-41.85,"Médio","Usa diesel — candidato solar/híbrido","Potência: 180 kW","","",""),
        ],
        "ES":[
            ("Rural-Irrigacao","Fazenda Caparaó","Alegre",160,-20.76,-41.53,"Médio","Verificar","Área: 160 ha","","",""),
            ("Rural-Irrigacao","Fazenda Norte Capixaba","São Mateus",220,-18.71,-39.86,"Médio","Sem rede próxima","Área: 220 ha","","",""),
            ("Rural-Irrigacao","Sítio Santa Maria","Linhares",95,-19.39,-40.06,"Médio","Verificar","Área: 95 ha","","",""),
            ("Mineracao","Mineração ES SA","Cachoeiro de Itapemirim",140,-20.84,-41.11,"Médio","Área remota — alto risco diesel","Concessão de Lavra · Mármore","","",""),
            ("Mineracao","Extratora Capixaba","Nova Venécia",95,-18.71,-40.40,"Médio","Área remota — alto risco diesel","Licenciamento · Granito","","",""),
            ("ACL-GrupoA","Aracruz Celulose SA","Aracruz",0,-19.82,-40.27,"Alto","Conectado — Grupo A / Mercado Livre","Demanda: 22.0 MW","67.890.123/0001-55","Carlos Lima","27991234567"),
            ("ACL-GrupoA","Siderúrgica Tubarão SA","Serra",0,-20.13,-40.30,"Alto","Conectado — Grupo A / Mercado Livre","Demanda: 45.0 MW","78.901.234/0001-66","Roberto Silva","27991234568"),
            ("Diesel-Gerador","Pedreiras Capixabas SA","Cachoeiro de Itapemirim",0,-20.84,-41.11,"Médio","Usa diesel — candidato solar/híbrido","Potência: 420 kW","","",""),
            ("Diesel-Gerador","Cerâmica Norte ES","São Mateus",0,-18.71,-39.86,"Médio","Usa diesel — candidato solar/híbrido","Potência: 190 kW","","",""),
        ],
    }
    demos = base.get(estado,[
        ("Rural-Irrigacao",f"Fazenda Demo {estado}","Capital",500,-15.0,-50.0,"Alto","Sem rede próxima","Área: 500 ha","","",""),
        ("Mineracao",f"Mineradora Demo {estado}","Interior",300,-15.5,-50.5,"Alto","Área remota — alto risco diesel","Concessão de Lavra","","",""),
        ("ACL-GrupoA",f"Indústria Demo {estado}","Capital",0,-15.0,-50.0,"Alto","Conectado — Grupo A / Mercado Livre","Demanda: 3.0 MW","","",""),
        ("Diesel-Gerador",f"Gerador Demo {estado}","Interior",0,-15.5,-50.5,"Médio","Usa diesel — candidato solar/híbrido","Potência: 400 kW","","",""),
    ])
    return [{"modulo":d[0],"nome":d[1],"municipio":d[2],"area_ha":d[3],"lat":d[4],"lon":d[5],"potencial":d[6],"situacao_rede":d[7],"observacoes":d[8],"estado":estado,"fonte":d[0].split("-")[0]+"-demo","cpf_cnpj":d[9],"socio":d[10],"telefone":d[11],"email":"","porte":"","cnae":"","codigo_ref":"DEMO","situacao":"Ativo"} for d in demos]

# ── APIs COM RETRY ────────────────────────────────────────────────────────────
HEADERS = {"User-Agent":"SolarLeadFinder/2.0","Accept":"application/json"}

def get_json_retry(url, params=None, timeout=35, tentativas=3):
    for i in range(tentativas):
        try:
            r = requests.get(url, params=params, headers=HEADERS, timeout=timeout)
            if r.status_code == 200: return r.json()
            if r.status_code == 429: time.sleep(15)
        except requests.exceptions.Timeout:
            time.sleep(3)
        except: pass
    return None

def buscar_sicar(estado, municipios_filtro=None):
    leads = []
    for pag in range(1, 8):
        d = get_json_retry("https://consultapublica.car.gov.br/publico/imoveis/index",
                           {"estado":estado,"tipo":"IRU","pagina":pag,"tamanhoPagina":100})
        if not d or not isinstance(d,list) or not d: break
        for item in d:
            area = float(item.get("area",0) or 0)
            if area < 50: continue
            mun = item.get("nomeMunicipio","")
            if municipios_filtro and normalizar(mun) not in [normalizar(m) for m in municipios_filtro]: continue
            lat = float(item.get("latitude") or 0)
            lon = float(item.get("longitude") or 0)
            if not lat or not lon:
                c = get_coords_municipio(mun)
                lat, lon = (c[0] or 0), (c[1] or 0)
            leads.append({"modulo":"Rural-Irrigacao","nome":item.get("nomePessoa",""),"municipio":mun,"area_ha":round(area,1),"potencial":"Alto" if area>=500 else "Médio","situacao_rede":"Verificar","observacoes":f"Área: {area:.0f} ha","lat":lat,"lon":lon,"fonte":"SICAR","cpf_cnpj":item.get("cpf",item.get("cnpj","")),"socio":"","telefone":"","email":"","porte":"","cnae":"","codigo_ref":item.get("codigo",""),"situacao":item.get("situacao","")})
        time.sleep(0.4)
    return leads

def buscar_anm(estado, municipios_filtro=None):
    leads = []
    for offset in range(0, 600, 100):
        d = get_json_retry("https://geo.anm.gov.br/geoserver/sigmine/wfs",
                           {"service":"WFS","version":"2.0.0","request":"GetFeature","typeName":"sigmine:BRASIL","outputFormat":"application/json","CQL_FILTER":f"UF='{estado}'","count":100,"startIndex":offset}, timeout=45)
        if not d or "features" not in d or not d["features"]: break
        for f in d["features"]:
            p = f.get("properties",{})
            area = float(p.get("AREA_HA",0) or 0)
            fase = p.get("FASE","")
            mun = p.get("MUNICIPIO","")
            if municipios_filtro and normalizar(mun) not in [normalizar(m) for m in municipios_filtro]: continue
            try:
                coords = f.get("geometry",{}).get("coordinates",[[[[0,0]]]])
                pts = coords[0][0] if f.get("geometry",{}).get("type")=="MultiPolygon" else coords[0]
                lat = round(sum(pt[1] for pt in pts)/len(pts),6)
                lon = round(sum(pt[0] for pt in pts)/len(pts),6)
            except:
                c = get_coords_municipio(mun)
                lat, lon = (c[0] or 0), (c[1] or 0)
            leads.append({"modulo":"Mineracao","nome":p.get("NOME",""),"municipio":mun,"area_ha":round(area,1),"potencial":"Alto" if area>=200 else "Médio","situacao_rede":"Área remota — alto risco diesel","observacoes":f"Fase: {fase} · Substância: {p.get('SUB','')}","lat":lat,"lon":lon,"fonte":"ANM-SIGMINE","cpf_cnpj":p.get("CPF_CNPJ",""),"socio":"","telefone":"","email":"","porte":"","cnae":"","codigo_ref":p.get("PROCESSO",""),"situacao":fase})
        time.sleep(0.5)
    return leads

def buscar_aneel(estado, tipo="diesel"):
    filtro = {"SigUF":estado}
    if tipo=="diesel": filtro["DscCombustivel"]="Óleo Diesel"
    d = get_json_retry("https://dadosabertos.aneel.gov.br/api/3/action/datastore_search",
                       {"resource_id":"b1bd71e7-d0ad-4214-9053-cbd58e9564a7","limit":500,"filters":json.dumps(filtro)})
    leads = []
    if d and d.get("success"):
        for r in d.get("result",{}).get("records",[]):
            pot = float(r.get("MdaPotenciaFiscalizadaKw",0) or 0)
            mun = r.get("NomMunicipio","")
            lat = float(r.get("NumLatitude") or 0)
            lon = float(r.get("NumLongitude") or 0)
            if not lat or not lon:
                c = get_coords_municipio(mun)
                lat, lon = (c[0] or 0), (c[1] or 0)
            if tipo=="diesel":
                leads.append({"modulo":"Diesel-Gerador","nome":r.get("NomEmpreendimento",""),"municipio":mun,"area_ha":0,"potencial":"Alto" if pot>=500 else "Médio","situacao_rede":"Usa diesel — candidato solar/híbrido","observacoes":f"Potência: {pot:.0f} kW","lat":lat,"lon":lon,"fonte":"ANEEL-BIG","cpf_cnpj":r.get("CpfCnpj",""),"socio":"","telefone":"","email":"","porte":"","cnae":"","codigo_ref":r.get("CodCEG",""),"situacao":r.get("DscFaseUsina","")})
            else:
                leads.append({"modulo":"ACL-GrupoA","nome":r.get("NomAgente",r.get("NomEmpreendimento","")),"municipio":mun,"area_ha":0,"potencial":"Alto","situacao_rede":"Conectado — Grupo A / Mercado Livre","observacoes":f"Categoria: {r.get('DscCategoria','')}","lat":lat,"lon":lon,"fonte":"ANEEL","cpf_cnpj":r.get("CnpjParticipante",r.get("CpfCnpj","")),"socio":"","telefone":"","email":"","porte":"","cnae":"","codigo_ref":r.get("CodCEG",""),"situacao":"Ativo"})
    return leads

# ── SCORE ─────────────────────────────────────────────────────────────────────
def calcular_score(lead):
    s = 5.0
    pot = lead.get("potencial","")
    if pot=="Alto": s+=2.5
    elif pot=="Médio": s+=1.0
    if "diesel" in str(lead.get("situacao_rede","")).lower(): s+=1.5
    if "remot" in str(lead.get("situacao_rede","")).lower(): s+=1.0
    if lead.get("modulo")=="ACL-GrupoA": s+=1.5
    area=float(lead.get("area_ha",0) or 0)
    if area>=500: s+=0.5
    elif area>=200: s+=0.3
    if lead.get("telefone"): s+=0.5
    if lead.get("cpf_cnpj") and "DEMO" not in str(lead.get("cpf_cnpj","")): s+=0.3
    dist=lead.get("dist_km")
    if dist and dist<=100: s+=0.3
    return round(min(s,10.0),1)

# ── BUSCA PRINCIPAL ────────────────────────────────────────────────────────────
def buscar_leads(estados, modulos, situacoes, score_min, cidade_ref, raio_km, regioes_sel):
    todos = []
    total_etapas = max(len(estados)*len(modulos),1)
    etapa = 0
    progress = st.progress(0)
    status = st.empty()
    fontes_reais = []

    # Geocodificar cidade de referência
    lat_ref, lon_ref = None, None
    if cidade_ref and cidade_ref.strip():
        lat_ref, lon_ref = get_coords_municipio(cidade_ref.strip())
        if lat_ref:
            status.markdown(f"<p style='color:#4ade80;font-size:12px'>📍 {cidade_ref} localizada ({lat_ref:.4f}, {lon_ref:.4f})</p>",unsafe_allow_html=True)
            time.sleep(0.5)
        else:
            status.markdown(f"<p style='color:#fbbf24;font-size:12px'>⚠ Cidade '{cidade_ref}' não encontrada — buscando em todo o estado</p>",unsafe_allow_html=True)
            time.sleep(1)

    for estado in estados:
        municipios_regiao = []
        reg = regioes_sel.get(estado,"Todas as regiões")
        if reg and reg != "Todas as regiões":
            municipios_regiao = REGIOES.get(estado,{}).get(reg,[])

        demos = gerar_demo(estado)

        for mod in modulos:
            etapa+=1
            progress.progress(int((etapa/total_etapas)*100))
            status.markdown(f"<p style='color:#94a3b8;font-size:12px'>Consultando {mod} em {estado}...</p>",unsafe_allow_html=True)
            leads_mod = []

            if mod=="Rural-Irrigacao":
                leads_mod = buscar_sicar(estado, municipios_regiao or None)
                if leads_mod: fontes_reais.append("SICAR")
                else: leads_mod = [l for l in demos if l["modulo"]=="Rural-Irrigacao"]

            elif mod=="Mineracao":
                leads_mod = buscar_anm(estado, municipios_regiao or None)
                if leads_mod: fontes_reais.append("ANM")
                else: leads_mod = [l for l in demos if l["modulo"]=="Mineracao"]

            elif mod=="ACL-GrupoA":
                leads_mod = buscar_aneel(estado, tipo="acl")
                if leads_mod: fontes_reais.append("ANEEL-ACL")
                else: leads_mod = [l for l in demos if l["modulo"]=="ACL-GrupoA"]

            elif mod=="Diesel-Gerador":
                leads_mod = buscar_aneel(estado, tipo="diesel")
                if leads_mod: fontes_reais.append("ANEEL-Diesel")
                else: leads_mod = [l for l in demos if l["modulo"]=="Diesel-Gerador"]

            # Coordenadas + distância
            for lead in leads_mod:
                lead["estado"] = estado
                if not lead.get("lat") or not lead.get("lon"):
                    c = get_coords_municipio(lead.get("municipio",""))
                    lead["lat"] = c[0] or 0
                    lead["lon"] = c[1] or 0
                if lat_ref and lon_ref and lead.get("lat") and lead.get("lon"):
                    lead["dist_km"] = round(distancia_km(lat_ref,lon_ref,lead["lat"],lead["lon"]),1)
                else:
                    lead["dist_km"] = None
                lead["score"] = calcular_score(lead)

            todos += leads_mod
            time.sleep(0.2)

    progress.empty()
    status.empty()

    if not todos: return pd.DataFrame(), fontes_reais

    df = pd.DataFrame(todos)

    # Filtro por raio
    if lat_ref and lon_ref and raio_km and raio_km>0:
        df = df[df["dist_km"].apply(lambda x: x is not None and x<=raio_km)]

    # Filtro por situação
    if situacoes and "Todos" not in situacoes:
        mask = pd.Series([False]*len(df), index=df.index)
        if "Zero-grid / Sem rede" in situacoes: mask |= df["situacao_rede"].str.contains("remot|Sem rede|zero",case=False,na=False)
        if "Usa diesel" in situacoes: mask |= df["situacao_rede"].str.contains("diesel",case=False,na=False)
        if "ACL / Grupo A" in situacoes: mask |= df["situacao_rede"].str.contains("Grupo A",case=False,na=False)
        if "Alto consumo" in situacoes: mask |= df["modulo"].isin(["ACL-GrupoA","Diesel-Gerador"])
        df = df[mask]

    df = df[df["score"]>=score_min]
    return df.sort_values("score",ascending=False).reset_index(drop=True), fontes_reais

# ── EXCEL ─────────────────────────────────────────────────────────────────────
def gerar_excel(df):
    colunas=["score","potencial","modulo","estado","municipio","nome","cpf_cnpj","socio","telefone","email","porte","area_ha","dist_km","situacao_rede","situacao","lat","lon","cnae","codigo_ref","observacoes","fonte"]
    for c in colunas:
        if c not in df.columns: df[c]=""
    df=df[colunas]
    output=io.BytesIO()
    with pd.ExcelWriter(output,engine="openpyxl") as writer:
        df.to_excel(writer,sheet_name="Todos os Leads",index=False)
        df[df["potencial"]=="Alto"].to_excel(writer,sheet_name="Alto Potencial",index=False)
        for mod in df["modulo"].unique():
            df[df["modulo"]==mod].to_excel(writer,sheet_name=str(mod)[:31].replace("/","-"),index=False)
        ws=writer.sheets["Todos os Leads"]
        from openpyxl.styles import PatternFill,Font,Alignment
        from openpyxl.utils import get_column_letter
        for cell in ws[1]:
            cell.fill=PatternFill("solid",fgColor="1C1C2E")
            cell.font=Font(color="F59E0B",bold=True,size=10)
            cell.alignment=Alignment(horizontal="center")
        cores={"Alto":("1A2E1A","4ADE80"),"Médio":("1A1A2E","60A5FA"),"Baixo":("2E1A1A","F87171")}
        for i,row in enumerate(ws.iter_rows(min_row=2),0):
            if i<len(df):
                bg,fc=cores.get(df.iloc[i].get("potencial",""),("1E293B","E2E8F0"))
                for cell in row:
                    cell.fill=PatternFill("solid",fgColor=bg)
                    cell.font=Font(color=fc,size=9)
        for i,w in enumerate([7,10,16,6,20,35,18,25,14,28,12,8,10,28,15,10,10,30,15,40,12],1):
            ws.column_dimensions[get_column_letter(i)].width=w
    return output.getvalue()

# ── INTERFACE ─────────────────────────────────────────────────────────────────
ch1,ch2=st.columns([5,1])
with ch1:
    st.markdown('<div style="display:flex;align-items:center;gap:12px;margin-bottom:8px"><div style="width:40px;height:40px;background:#f59e0b;border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:22px">☀️</div><div><h1 style="margin:0;font-size:22px;font-weight:500;color:#f1f5f9!important">Solar Lead Finder</h1><p style="margin:0;font-size:12px;color:#475569!important">Prospecção automática · Fontes públicas oficiais</p></div></div>',unsafe_allow_html=True)
with ch2:
    st.markdown("<br>",unsafe_allow_html=True)
    if st.button("Sair",key="logout"):
        st.session_state.autenticado=False
        st.rerun()

st.markdown('<span class="badge-src">SICAR</span><span class="badge-src">ANM</span><span class="badge-src">CCEE</span><span class="badge-src">ANEEL</span><span style="font-size:11px;color:#475569"> · Dados abertos do governo federal</span><hr style="margin:12px 0">',unsafe_allow_html=True)

with st.sidebar:
    st.markdown(f"<p style='font-size:12px;color:#475569'>Usuário: <b style='color:#f59e0b'>{st.session_state.get('usuario','')}</b></p>",unsafe_allow_html=True)
    st.markdown("### Filtros")
    st.markdown("---")

    estados_sel=st.multiselect("Estados",options=TODOS_ESTADOS,default=["MG","BA","ES"])

    regioes_sel={}
    if estados_sel:
        st.markdown("---")
        st.markdown("**Região (opcional)**")
        for est in estados_sel:
            if est in REGIOES:
                opcoes=[k for k in REGIOES[est].keys()]
                regioes_sel[est]=st.selectbox(f"Região de {est}",options=opcoes,key=f"reg_{est}")

    st.markdown("---")
    st.markdown("**Localização de referência**")
    cidade_ref=st.text_input("Cidade central",placeholder="Ex: Teófilo Otoni",help="Deixe vazio para buscar em todo o estado")
    raio_km=st.slider("Raio de busca (km)",0,500,0,25,help="0 = sem filtro de raio")
    if raio_km>0 and cidade_ref:
        st.caption(f"Raio de {raio_km} km ao redor de {cidade_ref}")

    st.markdown("---")
    st.markdown("**Módulos**")
    mod_rural =st.checkbox("Rural / Irrigação (SICAR)",value=True)
    mod_mine  =st.checkbox("Mineração (ANM)",value=True)
    mod_acl   =st.checkbox("ACL / Grupo A (CCEE)",value=True)
    mod_diesel=st.checkbox("Diesel / Gerador (ANEEL)",value=True)

    st.markdown("---")
    situacoes=st.multiselect("Situação energética",options=["Todos","Zero-grid / Sem rede","Usa diesel","ACL / Grupo A","Alto consumo"],default=["Todos"])

    st.markdown("---")
    score_min=st.slider("Score mínimo",0.0,10.0,6.0,0.5)

    st.markdown("---")
    buscar=st.button("Buscar leads agora",type="primary")

if "df_resultado" not in st.session_state:
    st.session_state.df_resultado=pd.DataFrame()
    st.session_state.buscou=False
    st.session_state.fontes_reais=[]

if buscar:
    if not estados_sel:
        st.error("Selecione pelo menos um estado.")
    else:
        modulos_sel=[]
        if mod_rural:  modulos_sel.append("Rural-Irrigacao")
        if mod_mine:   modulos_sel.append("Mineracao")
        if mod_acl:    modulos_sel.append("ACL-GrupoA")
        if mod_diesel: modulos_sel.append("Diesel-Gerador")
        if not modulos_sel:
            st.error("Selecione pelo menos um módulo.")
        else:
            with st.spinner("Consultando fontes públicas..."):
                df,fontes=buscar_leads(estados_sel,modulos_sel,situacoes,score_min,cidade_ref,raio_km if raio_km>0 else None,regioes_sel)
                st.session_state.df_resultado=df
                st.session_state.buscou=True
                st.session_state.fontes_reais=fontes

if st.session_state.buscou and not st.session_state.df_resultado.empty:
    df=st.session_state.df_resultado
    fontes_reais=st.session_state.fontes_reais

    ref_txt=f" · Raio: {raio_km} km de {cidade_ref}" if (raio_km and cidade_ref) else ""
    st.success(f"✓ {len(df)} leads encontrados{ref_txt}")

    # Mostrar quais fontes responderam de verdade
    if fontes_reais:
        fontes_str=" · ".join(set(fontes_reais))
        st.markdown(f'<p style="font-size:12px;color:#4ade80">✓ Fontes reais que responderam: {fontes_str}</p>',unsafe_allow_html=True)
    else:
        st.markdown('<p style="font-size:12px;color:#fbbf24">⚠ APIs com timeout — exibindo dados demonstrativos. Tente novamente em alguns minutos.</p>',unsafe_allow_html=True)

    c1,c2,c3,c4=st.columns(4)
    with c1: st.metric("Total de leads",len(df))
    with c2: st.metric("Alto potencial",len(df[df["potencial"]=="Alto"]))
    with c3: st.metric("Usam diesel",len(df[df["situacao_rede"].str.contains("diesel",case=False,na=False)]))
    with c4: st.metric("Com CNPJ",len(df[df["cpf_cnpj"].astype(str).str.len()>5]))

    st.markdown("---")
    cf1,cf2=st.columns([2,1])
    with cf1: filtro_mod=st.selectbox("Filtrar por módulo",["Todos"]+list(df["modulo"].unique()))
    with cf2: ordenar=st.selectbox("Ordenar por",["Score","Distância (km)","Município"])

    df_view=df if filtro_mod=="Todos" else df[df["modulo"]==filtro_mod]
    if ordenar=="Distância (km)" and "dist_km" in df_view.columns:
        df_view=df_view.sort_values("dist_km",na_position="last")
    elif ordenar=="Município":
        df_view=df_view.sort_values("municipio")

    cols=[c for c in ["score","potencial","estado","municipio","dist_km","nome","modulo","situacao_rede","telefone","observacoes","fonte"] if c in df_view.columns]
    st.dataframe(df_view[cols].reset_index(drop=True),use_container_width=True,height=440,
        column_config={
            "score":st.column_config.NumberColumn("Score",format="%.1f"),
            "dist_km":st.column_config.NumberColumn("Dist. km",format="%.0f km"),
            "estado":st.column_config.TextColumn("UF",width=55),
            "municipio":st.column_config.TextColumn("Município",width=140),
            "nome":st.column_config.TextColumn("Nome / Razão Social",width=210),
            "modulo":st.column_config.TextColumn("Módulo",width=120),
            "situacao_rede":st.column_config.TextColumn("Situação",width=180),
            "observacoes":st.column_config.TextColumn("Observações",width=200),
        })

    st.markdown("---")
    c1,c2=st.columns([2,1])
    with c1:
        ts=datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(label="Baixar planilha Excel completa",data=gerar_excel(df.copy()),file_name=f"leads_solar_{ts}.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with c2:
        st.caption(f"{len(df)} leads · {len(df[df['potencial']=='Alto'])} alto potencial")

elif st.session_state.buscou:
    st.warning("Nenhum lead encontrado. Reduza o score mínimo, aumente o raio ou amplie os estados.")
else:
    st.markdown('<div style="text-align:center;padding:60px 20px"><div style="font-size:52px;margin-bottom:16px">☀️</div><p style="font-size:16px;color:#64748b">Configure os filtros e clique em <b style="color:#f59e0b">Buscar leads agora</b></p><p style="font-size:13px;color:#374151;margin-top:8px">SICAR · ANM · CCEE · ANEEL · Receita Federal</p></div>',unsafe_allow_html=True)
